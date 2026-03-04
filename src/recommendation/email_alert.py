import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders
import os
import logging
import json
from pathlib import Path
from src.config.settings import POSTGRES

logger = logging.getLogger(__name__)
logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")

def get_engine():
    return create_engine(
        f"postgresql+psycopg2://{POSTGRES.user}:{POSTGRES.password}@{POSTGRES.host}/{POSTGRES.db}"
    )

def table_exists(engine, table_name: str) -> bool:
    """Return True if the given table exists in the database."""
    try:
        with engine.connect() as conn:
            res = conn.execute(f"SELECT to_regclass('{table_name}')").scalar()
            return res is not None
    except Exception:
        return False


def table_has_column(engine, table_name: str, column_name: str) -> bool:
    """Return True if the table contains the given column."""
    try:
        schema, table = table_name.split(".", 1) if "." in table_name else ("public", table_name)
        with engine.connect() as conn:
            res = conn.execute(
                text("""
                    SELECT EXISTS (
                        SELECT 1
                        FROM information_schema.columns
                        WHERE table_schema = :schema
                          AND table_name = :table
                          AND column_name = :column
                    )
                """),
                {"schema": schema, "table": table, "column": column_name},
            ).scalar()
            return bool(res)
    except Exception:
        return False


def load_walk_forward_validation_summary(tabs, reports_dir="/opt/airflow/files/reports"):
    """
    Build a flat summary DataFrame from walk-forward JSON outputs.
    One row per asset class + horizon.
    """
    def _summarize_regimes(regime_rows):
        if not regime_rows:
            return {}
        rdf = pd.DataFrame(regime_rows)
        if rdf.empty or "regime" not in rdf.columns:
            return {}
        required = {"samples", "hit_rate", "avg_return", "avg_benchmark_return"}
        if not required.issubset(set(rdf.columns)):
            return {}

        for col in ["samples", "hit_rate", "avg_return", "avg_benchmark_return"]:
            if col in rdf.columns:
                rdf[col] = pd.to_numeric(rdf[col], errors="coerce")

        out = {}
        for regime in ("bull", "bear", "sideways"):
            subset = rdf[rdf["regime"] == regime].copy()
            if subset.empty:
                continue

            weights = subset.get("samples", pd.Series(dtype=float)).fillna(0).clip(lower=0)
            weight_sum = float(weights.sum())
            if weight_sum > 0:
                hit_rate = float(np.average(subset["hit_rate"].fillna(0.0), weights=weights))
                avg_return = float(np.average(subset["avg_return"].fillna(0.0), weights=weights))
                avg_benchmark = float(np.average(subset["avg_benchmark_return"].fillna(0.0), weights=weights))
            else:
                hit_rate = float(subset["hit_rate"].mean(skipna=True))
                avg_return = float(subset["avg_return"].mean(skipna=True))
                avg_benchmark = float(subset["avg_benchmark_return"].mean(skipna=True))

            out[f"{regime}_samples"] = int(weights.sum())
            out[f"{regime}_hit_rate"] = hit_rate
            out[f"{regime}_avg_return"] = avg_return
            out[f"{regime}_avg_benchmark_return"] = avg_benchmark
        return out

    rows = []
    report_root = Path(reports_dir)
    for asset_class, (_, features_table) in tabs.items():
        safe_table = features_table.replace(".", "_")
        summary_path = report_root / f"{safe_table}_walk_forward_summary.json"
        if not summary_path.exists():
            rows.append(
                {
                    "asset_class": asset_class,
                    "source_table": features_table,
                    "horizon": "n/a",
                    "status": "missing_report",
                }
            )
            continue

        try:
            payload = json.loads(summary_path.read_text(encoding="utf-8"))
        except Exception as e:
            rows.append(
                {
                    "asset_class": asset_class,
                    "source_table": features_table,
                    "horizon": "n/a",
                    "status": "report_parse_failed",
                    "error": str(e),
                }
            )
            continue

        generated_utc = pd.Timestamp(summary_path.stat().st_mtime, unit="s", tz="UTC").strftime(
            "%Y-%m-%d %H:%M:%S UTC"
        )
        horizons = payload.get("horizons", [])
        if not horizons:
            rows.append(
                {
                    "asset_class": asset_class,
                    "source_table": features_table,
                    "horizon": "n/a",
                    "status": "no_horizons",
                    "report_generated_utc": generated_utc,
                }
            )
            continue

        for horizon_data in horizons:
            summary = horizon_data.get("summary", {})
            drift_summary = horizon_data.get("drift_summary", {}) or {}
            regime_summary = _summarize_regimes(horizon_data.get("regime_rows", []))
            rows.append(
                {
                    "asset_class": asset_class,
                    "source_table": features_table,
                    "horizon": horizon_data.get("horizon", "n/a"),
                    "status": horizon_data.get("status", "unknown"),
                    "selected_model_type": horizon_data.get("selected_model_type"),
                    "recommended_min_prob": horizon_data.get("recommended_min_prob"),
                    "splits_ran": summary.get("splits_ran"),
                    "avg_roc_auc": summary.get("avg_roc_auc"),
                    "avg_f1": summary.get("avg_f1"),
                    "avg_brier": summary.get("avg_brier"),
                    "avg_cagr_approx": summary.get("avg_cagr_approx"),
                    "avg_sharpe_annualized": summary.get("avg_sharpe_annualized"),
                    "avg_max_drawdown": summary.get("avg_max_drawdown"),
                    "avg_hit_rate": summary.get("avg_hit_rate"),
                    "drift_status": horizon_data.get("drift_status"),
                    "drift_avg_psi": drift_summary.get("avg_psi"),
                    "drift_avg_ks": drift_summary.get("avg_ks"),
                    "drift_high_features": drift_summary.get("high_drift_features"),
                    "drift_medium_features": drift_summary.get("medium_drift_features"),
                    "drift_recent_start": drift_summary.get("recent_start"),
                    "drift_recent_end": drift_summary.get("recent_end"),
                    "report_generated_utc": generated_utc,
                    **regime_summary,
                }
            )

    if not rows:
        return pd.DataFrame()

    out = pd.DataFrame(rows)
    preferred_order = [
        "asset_class",
        "source_table",
        "horizon",
        "status",
        "selected_model_type",
        "recommended_min_prob",
        "splits_ran",
        "avg_roc_auc",
        "avg_f1",
        "avg_brier",
        "avg_cagr_approx",
        "avg_sharpe_annualized",
        "avg_max_drawdown",
        "avg_hit_rate",
        "drift_status",
        "drift_avg_psi",
        "drift_avg_ks",
        "drift_high_features",
        "drift_medium_features",
        "drift_recent_start",
        "drift_recent_end",
        "bull_samples",
        "bull_hit_rate",
        "bull_avg_return",
        "bull_avg_benchmark_return",
        "bear_samples",
        "bear_hit_rate",
        "bear_avg_return",
        "bear_avg_benchmark_return",
        "sideways_samples",
        "sideways_hit_rate",
        "sideways_avg_return",
        "sideways_avg_benchmark_return",
        "report_generated_utc",
        "error",
    ]
    final_cols = [c for c in preferred_order if c in out.columns] + [c for c in out.columns if c not in preferred_order]
    return out[final_cols]


def build_validation_email_section(validation_df: pd.DataFrame) -> str:
    if validation_df.empty:
        return "Model Validation Snapshot: Not available."

    def _fmt_pct(value, decimals=1):
        return f"{float(value) * 100:.{decimals}f}%" if pd.notna(value) else "n/a"

    def _fmt_num(value, decimals=3):
        return f"{float(value):.{decimals}f}" if pd.notna(value) else "n/a"

    lines = ["Model Validation Snapshot (walk-forward):"]
    for _, row in validation_df.sort_values(["asset_class", "horizon"]).iterrows():
        asset_class = row.get("asset_class", "Unknown")
        horizon = row.get("horizon", "n/a")
        status = row.get("status", "unknown")
        if status != "ok":
            lines.append(f"- {asset_class} [{horizon}]: status={status}")
            continue

        model = row.get("selected_model_type", "n/a")
        threshold = row.get("recommended_min_prob")
        threshold_txt = _fmt_pct(threshold, decimals=0)
        roc_auc = row.get("avg_roc_auc")
        f1 = row.get("avg_f1")
        splits = row.get("splits_ran")
        roc_auc_txt = _fmt_num(roc_auc, decimals=3)
        f1_txt = _fmt_num(f1, decimals=3)
        splits_txt = str(int(splits)) if pd.notna(splits) else "n/a"
        drift_status = row.get("drift_status", "n/a")
        drift_high = row.get("drift_high_features")
        drift_medium = row.get("drift_medium_features")
        drift_txt = (
            f"{drift_status} (high={int(drift_high) if pd.notna(drift_high) else 'n/a'}, "
            f"medium={int(drift_medium) if pd.notna(drift_medium) else 'n/a'})"
        )
        regime_txt = (
            f"hit-rate bull/bear/sideways="
            f"{_fmt_pct(row.get('bull_hit_rate'))}/"
            f"{_fmt_pct(row.get('bear_hit_rate'))}/"
            f"{_fmt_pct(row.get('sideways_hit_rate'))}"
        )
        lines.append(
            f"- {asset_class} [{horizon}]: model={model}, threshold={threshold_txt}, "
            f"ROC-AUC={roc_auc_txt}, F1={f1_txt}, splits={splits_txt}, "
            f"drift={drift_txt}, {regime_txt}"
        )

    return "\n".join(lines)


def fetch_and_filter_data(engine, signals_table, features_table):
    # if either source or features table is missing, treat as empty
    if not table_exists(engine, signals_table):
        logger.warning(f"Signals table {signals_table} does not exist; skipping")
        return pd.DataFrame()
    if not table_exists(engine, features_table):
        logger.warning(f"Features table {features_table} does not exist; skipping")
        return pd.DataFrame()

    # signals_table and features_table should already include schema qualification
    has_asset_name = table_has_column(engine, features_table, "asset_name")
    asset_select = (
        "COALESCE(NULLIF(TRIM(f.asset_name), ''), s.symbol) AS asset_name,"
        if has_asset_name
        else "s.symbol AS asset_name,"
    )
    query = f"""
        SELECT s.*, {asset_select} f.close, f.rsi_14, f.volatility_7d, f.atr_14
        FROM (SELECT DISTINCT ON (symbol) * FROM {signals_table} ORDER BY symbol, trade_date DESC) s
        JOIN (SELECT DISTINCT ON (symbol) * FROM {features_table} ORDER BY symbol, event_time DESC) f
        ON s.symbol = f.symbol
    """
    try:
        df = pd.read_sql(query, engine)
    except Exception as e:
        logger.error(f"SQL error fetching data from {signals_table}/{features_table}: {e}")
        return pd.DataFrame()
    
    if df.empty:
        return pd.DataFrame()

    target_signals = ['INVEST NOW', 'ACCUMULATE', 'STRONG BUY', 'BUY']
    mask = pd.Series(False, index=df.index)
    
    if 'signal_5y' in df.columns:
        mask = mask | df['signal_5y'].isin(target_signals)
    if 'signal_1y' in df.columns:
        mask = mask | df['signal_1y'].isin(target_signals)
    if 'signal' in df.columns:
        mask = mask | df['signal'].isin(target_signals)
        
    filtered = df[mask].copy()
    if not filtered.empty:
        if 'risk_adjusted_score' in filtered.columns:
            filtered = filtered.sort_values('risk_adjusted_score', ascending=False, na_position='last')
        elif 'confidence_1y' in filtered.columns:
            filtered = filtered.sort_values('confidence_1y', ascending=False, na_position='last')
        elif 'confidence' in filtered.columns:
            filtered = filtered.sort_values('confidence', ascending=False, na_position='last')
    
    # Organize columns
    cols_to_keep = ['asset_name']
    
    if 'signal_1y' in df.columns and 'confidence_1y' in df.columns:
        cols_to_keep.extend(['signal_1y', 'confidence_1y'])
    if 'signal_5y' in df.columns and 'confidence_5y' in df.columns:
        cols_to_keep.extend(['signal_5y', 'confidence_5y'])
    if 'signal' in df.columns and 'confidence' in df.columns:
        cols_to_keep.extend(['signal', 'confidence'])

    for risk_col in [
        'combined_confidence',
        'risk_bucket',
        'risk_score',
        'suggested_position_pct',
        'expected_return_1y',
        'risk_adjusted_score',
        'var_95_1d',
        'cvar_95_1d',
    ]:
        if risk_col in df.columns:
            cols_to_keep.append(risk_col)

    cols_to_keep.extend(['close', 'rsi_14', 'volatility_7d', 'atr_14'])
    
    available_cols = [c for c in cols_to_keep if c in filtered.columns]
    return filtered[available_cols] if not filtered.empty else pd.DataFrame()


def write_excel_report(engine, tabs, filepath, validation_df=None):
    validation_df = validation_df if validation_df is not None else load_walk_forward_validation_summary(tabs)
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for tab_name, (sig_tbl, feat_tbl) in tabs.items():
            try:
                df = fetch_and_filter_data(engine, sig_tbl, feat_tbl)
                if not df.empty:
                    df.to_excel(writer, sheet_name=tab_name, index=False)
                else:
                    pd.DataFrame({'Message': ['No assets matched INVEST NOW/ACCUMULATE']}).to_excel(writer, sheet_name=tab_name, index=False)
            except Exception as e:
                logger.error(f"Error processing {tab_name}: {e}")
                pd.DataFrame({'Error': [str(e)]}).to_excel(writer, sheet_name=tab_name, index=False)

        if validation_df.empty:
            pd.DataFrame(
                {"Message": ["No validation summary available. Run ai_quant_model_validation_weekly."]}
            ).to_excel(writer, sheet_name="Validation Summary", index=False)
        else:
            validation_df.to_excel(writer, sheet_name="Validation Summary", index=False)

    # after writing, apply conditional formatting based on signal columns
    from openpyxl.styles import PatternFill
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl import load_workbook

    wb = load_workbook(filepath)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        # find signal columns used to trigger row highlight
        signal_col_letters = []
        confidence_col_letters = []
        for colcell in ws[1]:
            if colcell.value in {'signal', 'signal_1y', 'signal_5y'}:
                signal_col_letters.append(colcell.column_letter)
            if colcell.value in {'confidence', 'confidence_1y', 'confidence_5y', 'combined_confidence', 'suggested_position_pct', 'expected_return_1y', 'var_95_1d', 'cvar_95_1d'}:
                confidence_col_letters.append(colcell.column_letter)
        if signal_col_letters and ws.max_row >= 2:
            # apply fill to entire row when any signal column is INVEST NOW
            fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')
            formula_parts = [f'UPPER(${col}2)="INVEST NOW"' for col in signal_col_letters]
            formula = f"OR({','.join(formula_parts)})"
            # apply rule across all columns A..last
            last_col = ws.max_column
            last_letter = ws.cell(row=1, column=last_col).column_letter
            ws.conditional_formatting.add(
                f"A2:{last_letter}{ws.max_row}",
                FormulaRule(formula=[formula], stopIfTrue=True, fill=fill),
            )
        if confidence_col_letters and ws.max_row >= 2:
            # display confidence values as percentages (e.g., 0.58 -> 58.00%)
            for col in confidence_col_letters:
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws[f"{col}{row_idx}"]
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = "0.00%"
        if ws.max_row >= 2:
            for colcell in ws[1]:
                if colcell.value in {
                    "recommended_min_prob",
                    "avg_hit_rate",
                    "avg_cagr_approx",
                    "avg_max_drawdown",
                    "bull_hit_rate",
                    "bear_hit_rate",
                    "sideways_hit_rate",
                    "bull_avg_return",
                    "bear_avg_return",
                    "sideways_avg_return",
                    "bull_avg_benchmark_return",
                    "bear_avg_benchmark_return",
                    "sideways_avg_benchmark_return",
                }:
                    col = colcell.column_letter
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f"{col}{row_idx}"]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.00%"
                if colcell.value in {"risk_score"}:
                    col = colcell.column_letter
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws[f"{col}{row_idx}"]
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "0.0"
    wb.save(filepath)

def send_opportunity_email():
    engine = get_engine()
    
    # Definition of tabs - use medallion schema names, not public
    tabs = {
        "Nifty 50": ("gold.nifty50_investment_signals", "silver.nifty50_features_daily"),
        "Nifty Mid Cap": ("gold.nifty_midcap_investment_signals", "silver.nifty_midcap_features_daily"),
        "Nifty Small Cap": ("gold.nifty_smallcap_investment_signals", "silver.nifty_smallcap_features_daily"),
        "Mutual Funds": ("gold.mutual_funds_investment_signals", "silver.mutual_funds_features_daily"),
        "Crypto": ("gold.crypto_investment_signals", "silver.crypto_features_daily")
    }
    
    filepath = "/opt/airflow/files/investment_opportunities.xlsx"
    os.makedirs(os.path.dirname(filepath), exist_ok=True)
    validation_df = load_walk_forward_validation_summary(tabs)
    validation_snapshot = build_validation_email_section(validation_df)

    try:
        write_excel_report(engine, tabs, filepath, validation_df=validation_df)
    except PermissionError:
        fallback_filepath = f"/opt/airflow/files/investment_opportunities_{pd.Timestamp.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        logger.warning(f"Permission denied for {filepath}; retrying with {fallback_filepath}")
        try:
            write_excel_report(engine, tabs, fallback_filepath, validation_df=validation_df)
            filepath = fallback_filepath
        except Exception as e:
            logger.error(f"Failed to create Excel file after fallback: {e}")
            engine.dispose()
            return
    except Exception as e:
        logger.error(f"Failed to create Excel file: {e}")
        engine.dispose()
        return
    
    # Read SMTP and email configuration
    smtp_server = os.environ.get("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.environ.get("SMTP_PORT", "587"))
    smtp_user = os.environ.get("SMTP_USER", "")
    smtp_pass = os.environ.get("SMTP_PASSWORD", "")
    sender = os.environ.get("SENDER_EMAIL", "noreply@investmentplatform.local")
    receiver = os.environ.get("RECEIVER_EMAIL")

    # allow Airflow Variables to override or supply missing values
    try:
        from airflow.models import Variable
        if not smtp_user:
            smtp_user = Variable.get("SMTP_USER", default_var="")
        if not smtp_pass:
            smtp_pass = Variable.get("SMTP_PASSWORD", default_var="")
        if not receiver:
            receiver = Variable.get("RECEIVER_EMAIL", default_var="user@example.com")
        if not sender:
            sender = Variable.get("SENDER_EMAIL", default_var="noreply@investmentplatform.local")
    except ImportError:
        # running outside Airflow environment, ignore
        pass

    receiver = receiver or "user@example.com"
    recipients = [
        addr.strip()
        for addr in receiver.replace(";", ",").split(",")
        if addr.strip()
    ]
    if not recipients:
        recipients = ["user@example.com"]
    receiver_display = ", ".join(recipients)

    if not smtp_user or not smtp_pass:
        logger.warning("SMTP credentials not provided in env or Airflow Variables. Skipping email dispatch, but file generated.")
        return
        
    msg = MIMEMultipart()
    msg['From'] = sender
    msg['To'] = receiver_display
    msg['Subject'] = "Daily AI Investment Opportunities | Crypto & Nifty Market Signals"
    
    body = f"""Attached is the latest daily report for assets marked 'INVEST NOW' or 'ACCUMULATE'.

{validation_snapshot}

Disclaimer: This is not investment advice and is created for educational purposes only. Please consult a qualified financial advisor before making any investment decisions. Use this information at your own risk.

- AI Investment Engine"""

    msg.attach(MIMEText(body, 'plain'))
    
    with open(filepath, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", f"attachment; filename=investment_opportunities.xlsx")
    msg.attach(part)
    
    try:
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(smtp_user, smtp_pass)
        server.sendmail(sender, recipients, msg.as_string())
        server.quit()
        logger.info(f"Email sent successfully to {receiver_display}.")
    except Exception as e:
        logger.error(f"Failed to send email: {e}")
    finally:
        engine.dispose()

if __name__ == '__main__':
    send_opportunity_email()
